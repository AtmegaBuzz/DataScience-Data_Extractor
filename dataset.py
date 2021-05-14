from openpyxl import Workbook, load_workbook
import numpy as np

print("""
        Data Extractor is a python tool which allow data from exel sheets
        to be converted to a pandas dataframe or Series. Used for Data 
        Science and Data Gathering

        If you dont have openpyxl installed just install python and type in
        Cmd or Terminal
        pip install openpyxl
            or
        pip3 install openpyxl
        """,end = "\n")


class Extractor:

    def __init__(self,workbook_location,worksheet_location,column_name,start,end):
        self.workbook_location = workbook_location
        self.worksheet_location = worksheet_location
        self.column_name = column_name
        self.start = start
        self.end = end

    @staticmethod
    def extract(ws,column_name,start,end):
            data_set = []
            for i in range(start,end+1):
                data_set.append(ws[str(column_name+str(i))].value)

            return data_set


    def run_Extractor(self):
    
        
            #accessing workbook and worksheets
        my_workbook = load_workbook("{}".format(self.workbook_location))
        
            
        my_worksheet = my_workbook["{}".format(self.worksheet_location)]

            #list of the values int he column from given range
        data_Lst = self.extract(my_worksheet,self.column_name,self.start,self.end)
            
            
       

        return data_Lst


    

# location = str(input("Absolute Location Workbook: "))
# print(load_workbook(location).sheetnames)

# worksheet = str(input("Sheet-Name: "))

# column_name = str(input("Column: [example: A or B or C..]  "))
# start = int(input("Row Start Number[Look in your Exel]: "))
# end = int(input("Row End Number[Exclude column name]: "))

# extractor_bot = Extractor(location,worksheet,column_name,start,end)

# data = np.array(extractor_bot.run_Extractor())

# print(data)